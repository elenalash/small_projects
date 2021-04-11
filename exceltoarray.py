from pathlib import Path
from pandas import read_excel
from openpyxl import load_workbook

my_path = Path('C:/Users/Elena.Lashkova/Downloads/australia-unemployment-rate.xlsx')

# my_dataframe = read_excel(my_path,index_col=0)

wb = load_workbook(filename=my_path)

ws = wb.active

my_list = []

for i in range(18,47):
    active_row = ws[i]
    year = active_row[0].value.year
    percent = active_row[1].value
    my_list.append([year, percent])
    

print(my_list)